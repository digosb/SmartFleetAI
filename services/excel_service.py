import os
from openpyxl import Workbook, load_workbook

from services.config_service import ConfigService


class ExcelService:

    def __init__(self):

        self.config = ConfigService()

        self.headers = [
            "NOME",
            "DATA",
            "KM SAIDA",
            "HORA SAIDA",
            "KM CHEGADA",
            "HORA CHEGADA",
            "DESTINO",
            "CARRO",
            "PLACA"
        ]

    def carregar_planilha(self):

        caminho = self.config.obter_caminho_salvo()

        if not caminho:
            return None

        if not os.path.exists(caminho):

            workbook = Workbook()

            worksheet = workbook.active

            worksheet.title = "Controle dos Veiculos"

            worksheet.append(self.headers)

            workbook.save(caminho)

        return load_workbook(caminho)
    
    def salvar_viagem(self, dados):

        caminho = self.config.obter_caminho_salvo()

        if not caminho:
            return False

        workbook = self.carregar_planilha()

        if workbook is None:
            return False

        worksheet = workbook["Controle dos Veiculos"]

        worksheet.append([
            dados["nome"],
            dados["data"],
            dados["km_saida"],
            dados["hora_saida"],
            dados["km_chegada"],
            dados["hora_chegada"],
            dados["destino"],
            dados["carro"],
            dados["placa"]
        ])

        workbook.save(caminho)

        return True