import os # Bliblioteca de Sistema Operacional, serve para navegar dentro do computador, acessar pastas, arquivos. etc.
import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import load_workbook, Workbook

EXCEL_FILE = os.path.join(os.path.dirname(__file__), "CONTROLE KM CARRO.xlsx")
HEADERS = ["NOME", "DATA", "KM SAIDA", "HORA SAIDA", "KM CHEGADA", "HORA CHEGADA", "DESTINO", "CARRO", "PLACA"]


def carregar_planilha():
    if not os.path.exists(EXCEL_FILE):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Controle dos Veiculos"
        worksheet.append(HEADERS)
        workbook.save(EXCEL_FILE)
        return workbook

    workbook = load_workbook(EXCEL_FILE)
    if "Controle dos Veiculos" in workbook.sheetnames:
        worksheet = workbook["Controle dos Veiculos"]
    else:
        worksheet = workbook.create_sheet("Controle dos Veiculos")
        worksheet.append(HEADERS)
        workbook.save(EXCEL_FILE)
        return workbook

    if worksheet.max_row == 1 and all(cell.value != HEADERS[idx] for idx, cell in enumerate(worksheet[1])):
        worksheet.delete_rows(1, worksheet.max_row)
        worksheet.append(HEADERS)
        workbook.save(EXCEL_FILE)
    return workbook


def obter_ultima_linha(worksheet):
    linha = worksheet.max_row
    while linha > 1:
        if any(cell.value not in (None, "") for cell in worksheet[linha]):
            return linha
        linha -= 1
    return 1


def salvar_dados():
    dados = {
        "nome": entry_nome.get(),
        "data": entry_data.get(),
        "km_saida": entry_km_saida.get(),
        "hora_saida": entry_hora_saida.get(),
        "km_chegada": entry_km_chegada.get(),
        "hora_chegada": entry_hora_chegada.get(),
        "destino": entry_destino.get(),
        "carro": entry_carro.get(),
        "placa": entry_placa.get(),
    }

    if not dados["nome"] or not dados["data"]:
        messagebox.showwarning("Atenção", "Por favor, preencha pelo menos Nome e Data.")
        return

    workbook = carregar_planilha()
    worksheet = workbook.active
    ultima_linha = obter_ultima_linha(worksheet)
    nova_linha = ultima_linha + 1 if ultima_linha >= 1 else 2

    worksheet.append([
        dados["nome"],
        dados["data"],
        dados["km_saida"],
        dados["hora_saida"],
        dados["km_chegada"],
        dados["hora_chegada"],
        dados["destino"],
        dados["carro"],
        dados["placa"],
    ])
    workbook.save(EXCEL_FILE)

    print(f"Linha {nova_linha} gravada em {EXCEL_FILE}:")
    for chave, valor in dados.items():
        print(f"{chave}: {valor}")

    messagebox.showinfo("Sucesso", "Dados salvos na planilha com sucesso!")
    limpar_campos()


def limpar_campos():
    for campo in campos:
        campo.delete(0, tk.END)


root = tk.Tk()
root.title("Registro de Viagem")
root.geometry("520x430")
root.resizable(False, False)

frame = ttk.Frame(root, padding=20)
frame.pack(fill="both", expand=True)

ttk.Label(frame, text="Cadastro de Viagem", font=("Arial", 14, "bold")).grid(row=0, column=0, columnspan=2, pady=(0, 15))

campos = []

labels = [
    ("Nome", 1),
    ("Data", 2),
    ("KM Saída", 3),
    ("Hora Saída", 4),
    ("KM Chegada", 5),
    ("Hora Chegada", 6),
    ("Destino", 7),
    ("Carro", 8),
    ("Placa", 9),
]

for texto, linha in labels:
    ttk.Label(frame, text=texto).grid(row=linha, column=0, sticky="w", pady=4)
    entry = ttk.Entry(frame, width=30)
    entry.grid(row=linha, column=1, pady=4)
    campos.append(entry)

entry_nome = campos[0]
entry_data = campos[1]
entry_km_saida = campos[2]
entry_hora_saida = campos[3]
entry_km_chegada = campos[4]
entry_hora_chegada = campos[5]
entry_destino = campos[6]
entry_carro = campos[7]
entry_placa = campos[8]

buttons_frame = ttk.Frame(frame)
buttons_frame.grid(row=10, column=0, columnspan=2, pady=15)

btn_salvar = ttk.Button(buttons_frame, text="Salvar", command=salvar_dados)
btn_salvar.pack(side="left", padx=5)

btn_limpar = ttk.Button(buttons_frame, text="Limpar", command=limpar_campos)
btn_limpar.pack(side="left", padx=5)

root.mainloop()

